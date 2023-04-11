import json
import random
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .utils import create_excel_file
import random
import os
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
from calendar import monthrange

# Create your views here.
def home(request):
    return render(request, 'home.html')

@csrf_exempt
def generate_menu(request):
    if request.method == 'POST':
        # 获取表单数据
        data = json.loads(request.body)
        
        ingredients = data['ingredients']
        month = int(data['month'])
        breakfast_count = int(data['breakfastCount'])
        lunch_count = int(data['lunchCount'])
        dinner_count = int(data['dinnerCount'])

        # 计算当月天数
        now = datetime.now()
        days_in_month = monthrange(now.year, month)[1]

        # 生成菜单
        menu = ""
        for day in range(1, days_in_month + 1):
            menu += f"{month}月{day}日<br>早餐：<br>"
            menu += ", ".join(random.sample(ingredients, breakfast_count)) + "<br>午餐：<br>"
            menu += ", ".join(random.sample(ingredients, lunch_count)) + "<br>晚餐：<br>"
            menu += ", ".join(random.sample(ingredients, dinner_count)) + "<br><br>"

        # 创建并保存Excel文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "菜单"
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.cell(row=1, column=1, value="日期")
        ws.cell(row=1, column=2, value="早餐")
        ws.cell(row=1, column=3, value="午餐")
        ws.cell(row=1, column=4, value="晚餐")

        for day in range(1, days_in_month + 1):
            ws.cell(row=day + 1, column=1, value=f"{month}月{day}日")
            ws.cell(row=day + 1, column=2, value=", ".join(random.sample(ingredients, breakfast_count)))
            ws.cell(row=day + 1, column=3, value=", ".join(random.sample(ingredients, lunch_count)))
            ws.cell(row=day + 1, column=4, value=", ".join(random.sample(ingredients, dinner_count)))

        filename = f"menu_{month}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('app', 'generated_files', filename)
        wb.save(filepath)

        return JsonResponse({'menu': menu, 'filename': filename})
    else:
        return HttpResponse(status=405)

def download_menu(request, filename):
    filepath = os.path.join('app', 'generated_files', filename)
    if os.path.exists(filepath):
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={filename}'
            return response
    else:
        return HttpResponse(status=404)

def menu_generator(request):
    return render(request, 'app/menu_generator.html')